"""CSV/XLSX 导入解析、安全限制与幂等键覆盖测试。"""

from __future__ import annotations

import io
import zipfile
from collections.abc import Iterator
from typing import Any

import pytest
from openpyxl import Workbook

from app.errors import BusinessRuleError
from app.finance.importers import tabular
from app.finance.importers.tabular import parse_transaction_file


def test_csv_import_is_stable_across_retries_and_preserves_duplicate_rows() -> None:
    content = (
        "transaction_date,transaction_type,amount,category,currency,description\n"
        "2026-07-01,expense,25.50,餐饮,CNY,午餐\n"
        "2026-07-01,expense,25.50,餐饮,CNY,午餐\n"
    ).encode()

    first = parse_transaction_file("transactions.csv", content)
    retried = parse_transaction_file("transactions.csv", content)

    assert [item.import_key for item in first] == [item.import_key for item in retried]
    assert first[0].import_key != first[1].import_key
    assert first[0].row_number == 2


def test_external_id_is_stable_when_file_layout_changes() -> None:
    first = (
        b"transaction_date,transaction_type,amount,category,external_id\n"
        b"2026-07-01,income,100,salary,bank-42\n"
    )
    second = (
        b"transaction_date,transaction_type,amount,category,external_id\n"
        b"2026-07-02,income,200,bonus,bank-42\n"
    )

    assert parse_transaction_file("a.csv", first)[0].import_key == parse_transaction_file(
        "b.csv", second
    )[0].import_key


def test_xlsx_import_reads_values_without_evaluating_formulas() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        ["transaction_date", "transaction_type", "amount", "category", "currency"]
    )
    worksheet.append(["2026-07-01", "expense", "88.8", "交通", "CNY"])
    content = io.BytesIO()
    workbook.save(content)
    workbook.close()

    rows = parse_transaction_file("transactions.xlsx", content.getvalue())

    assert len(rows) == 1
    assert rows[0].values["category"] == "交通"


def test_import_rejects_unknown_or_missing_columns() -> None:
    content = (
        b"transaction_date,transaction_type,amount,unexpected\n"
        b"2026-07-01,expense,25,value\n"
    )

    with pytest.raises(BusinessRuleError, match="missing required columns"):
        parse_transaction_file("transactions.csv", content)


def _compressed_archive(member_name: str, payload: bytes) -> bytes:
    content = io.BytesIO()
    with zipfile.ZipFile(content, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(member_name, payload)
    return content.getvalue()


def test_xlsx_rejects_excessive_expanded_archive_size(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(tabular, "MAX_XLSX_UNCOMPRESSED_BYTES", 1024)
    monkeypatch.setattr(tabular, "MAX_XLSX_MEMBER_BYTES", 4096)
    monkeypatch.setattr(tabular, "MIN_XLSX_RATIO_CHECK_BYTES", 10_000)
    content = _compressed_archive("xl/worksheets/sheet1.xml", b"0" * 2048)

    with pytest.raises(BusinessRuleError, match="total expanded size limit"):
        parse_transaction_file("compressed-bomb.xlsx", content)


def test_xlsx_rejects_unsafe_compression_ratio(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(tabular, "MAX_XLSX_UNCOMPRESSED_BYTES", 10_000)
    monkeypatch.setattr(tabular, "MAX_XLSX_MEMBER_BYTES", 10_000)
    monkeypatch.setattr(tabular, "MIN_XLSX_RATIO_CHECK_BYTES", 100)
    monkeypatch.setattr(tabular, "MAX_XLSX_COMPRESSION_RATIO", 2)
    content = _compressed_archive("xl/sharedStrings.xml", b"0" * 4096)

    with pytest.raises(BusinessRuleError, match="compression ratio is unsafe"):
        parse_transaction_file("compressed-bomb.xlsx", content)


def test_xlsx_rejects_excessive_member_count_before_opening_zip(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    end_record = tabular.ZIP_EOCD_STRUCT.pack(
        tabular.ZIP_EOCD_SIGNATURE,
        0,
        0,
        tabular.MAX_XLSX_ARCHIVE_MEMBERS + 1,
        tabular.MAX_XLSX_ARCHIVE_MEMBERS + 1,
        0,
        0,
        0,
    )

    def fail_if_opened(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("ZIP parser should not be reached")

    monkeypatch.setattr(tabular.zipfile, "ZipFile", fail_if_opened)

    with pytest.raises(BusinessRuleError, match="too many members"):
        parse_transaction_file("member-bomb.xlsx", end_record)


class StreamingWorksheet:
    def __init__(self) -> None:
        self.yielded_data_rows = 0

    def iter_rows(self, *, values_only: bool) -> Iterator[tuple[Any, ...]]:
        assert values_only is True
        yield ("transaction_date", "transaction_type", "amount", "category")
        while True:
            self.yielded_data_rows += 1
            yield ("2026-07-01", "expense", "1", "测试")


class StreamingWorkbook:
    def __init__(self, worksheet: StreamingWorksheet) -> None:
        self.active = worksheet
        self.closed = False

    def close(self) -> None:
        self.closed = True


def test_xlsx_row_limit_stops_iterator_immediately(monkeypatch: pytest.MonkeyPatch) -> None:
    worksheet = StreamingWorksheet()
    workbook = StreamingWorkbook(worksheet)
    monkeypatch.setattr(tabular, "MAX_IMPORT_ROWS", 3)
    monkeypatch.setattr(tabular, "_validate_xlsx_archive", lambda _content: None)
    monkeypatch.setattr(tabular, "load_workbook", lambda *_args, **_kwargs: workbook)

    with pytest.raises(BusinessRuleError, match="3 row limit"):
        tabular._read_xlsx(b"bounded-stream")

    assert worksheet.yielded_data_rows == 4
    assert workbook.closed is True
