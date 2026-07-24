"""安全、确定性地解析 CSV 和 XLSX 交易导入文件。"""

# 本模块只提取文本值，领域校验在后续步骤执行。
from __future__ import annotations

import csv
import hashlib
import io
import struct
import zipfile
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from pathlib import PurePath, PurePosixPath
from typing import Any, Protocol, cast

from openpyxl import load_workbook

from app.errors import BusinessRuleError

# 解析刻意保持本地和确定性，不调用表格宏、外部链接或公式引擎。
# 在业务校验前执行限制，以控制内存和 CPU 消耗。
# 表头白名单可防止电子表格列变成任意字段。
MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
MAX_XLSX_ARCHIVE_MEMBERS = 256
MAX_XLSX_UNCOMPRESSED_BYTES = 32 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 16 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100
MIN_XLSX_RATIO_CHECK_BYTES = 1 * 1024 * 1024
XLSX_VALIDATION_CHUNK_BYTES = 64 * 1024
ZIP_EOCD_SIGNATURE = b"PK\x05\x06"
ZIP_EOCD_STRUCT = struct.Struct("<4s4H2LH")
REQUIRED_COLUMNS = {
    "transaction_date",
    "transaction_type",
    "amount",
    "category",
}
OPTIONAL_COLUMNS = {"currency", "description", "external_id"}
SUPPORTED_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS


class WorksheetReader(Protocol):
    """普通工作簿与只读工作簿共享的最小工作表接口。"""

    def iter_rows(self, *, values_only: bool) -> Iterator[tuple[Any, ...]]: ...


@dataclass(frozen=True, slots=True)
class ParsedTransactionRow:
    """领域校验前的源数据行及其确定性重试标识。"""

    row_number: int
    values: dict[str, Any]
    import_key: str


def parse_transaction_file(filename: str, content: bytes) -> list[ParsedTransactionRow]:
    """在不执行公式、不信任文件名的前提下解析上传表格。"""

    # 传给格式专用安全读取器的是文件内容，而不是 MIME 元数据。
    if not content:
        raise BusinessRuleError("import file is empty")
    if len(content) > MAX_IMPORT_BYTES:
        raise BusinessRuleError("import file exceeds the 10 MiB limit")

    # 后缀只用于选择解析器，两种解析器仍会分别校验文件结构。
    suffix = PurePath(filename).suffix.casefold()
    # 只接受文档中规定的两种不可执行表格格式。
    if suffix == ".csv":
        rows = _read_csv(content)
    elif suffix == ".xlsx":
        rows = _read_xlsx(content)
    else:
        raise BusinessRuleError("only CSV and XLSX transaction files are supported")
    return _normalize_rows(rows, hashlib.sha256(content).hexdigest())


def _read_csv(content: bytes) -> list[dict[str, Any]]:
    """解码 UTF-8 CSV，并保留每个源数据行以供校验。"""

    # 使用允许 BOM 的 UTF-8，可避免依赖平台的编码猜测
    # 在重试之间改变分类文本或幂等行为。
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise BusinessRuleError("CSV files must use UTF-8 encoding") from exc

    # `DictReader` 将首行视为字段名，且绝不执行单元格内容。
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise BusinessRuleError("import file must contain a header row")
    _validate_headers(reader.fieldnames)
    rows: list[dict[str, Any]] = []
    for row in reader:
        if len(rows) >= MAX_IMPORT_ROWS:
            raise BusinessRuleError(f"import file exceeds the {MAX_IMPORT_ROWS} row limit")
        rows.append(dict(row))
    return rows


def _xlsx_declared_member_count(content: bytes) -> int:
    """在构造 ZipInfo 列表前读取 EOCD，阻断超大中央目录。"""

    minimum_size = ZIP_EOCD_STRUCT.size
    search_start = max(0, len(content) - minimum_size - 65535)
    offset = content.rfind(ZIP_EOCD_SIGNATURE, search_start)
    if offset < 0 or len(content) - offset < minimum_size:
        raise BusinessRuleError("XLSX file is not a valid ZIP archive")

    (
        _signature,
        disk_number,
        central_directory_disk,
        members_on_disk,
        member_count,
        central_directory_size,
        central_directory_offset,
        comment_length,
    ) = ZIP_EOCD_STRUCT.unpack_from(content, offset)
    if offset + minimum_size + comment_length != len(content):
        raise BusinessRuleError("XLSX ZIP archive has an invalid end record")
    if (
        disk_number != 0
        or central_directory_disk != 0
        or members_on_disk != member_count
    ):
        raise BusinessRuleError("multi-disk XLSX archives are not supported")
    if (
        member_count == 0xFFFF
        or central_directory_size == 0xFFFFFFFF
        or central_directory_offset == 0xFFFFFFFF
    ):
        raise BusinessRuleError("ZIP64 XLSX archives are not supported")
    if member_count > MAX_XLSX_ARCHIVE_MEMBERS:
        raise BusinessRuleError("XLSX archive contains too many members")
    if central_directory_offset + central_directory_size > offset:
        raise BusinessRuleError("XLSX ZIP archive has an invalid central directory")
    return int(member_count)


def _validate_xlsx_member(member: zipfile.ZipInfo) -> None:
    """拒绝路径、加密、压缩算法或声明大小不安全的归档成员。"""

    path = PurePosixPath(member.filename)
    if (
        not member.filename
        or "\\" in member.filename
        or path.is_absolute()
        or ".." in path.parts
    ):
        raise BusinessRuleError("XLSX archive contains an unsafe member path")
    if member.flag_bits & 0x1:
        raise BusinessRuleError("encrypted XLSX archive members are not supported")
    if member.compress_type not in {zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED}:
        raise BusinessRuleError("XLSX archive uses an unsupported compression method")
    if member.file_size > MAX_XLSX_MEMBER_BYTES:
        raise BusinessRuleError("XLSX archive member exceeds the expanded size limit")


def _validate_xlsx_archive(content: bytes) -> None:
    """以有界流式解压验证 XLSX 容器，再允许 XML 解析。"""

    declared_member_count = _xlsx_declared_member_count(content)
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) != declared_member_count:
                raise BusinessRuleError("XLSX ZIP member count is inconsistent")

            total_uncompressed = 0
            total_compressed = 0
            for member in members:
                _validate_xlsx_member(member)
                if member.is_dir():
                    continue

                member_uncompressed = 0
                with archive.open(member) as source:
                    while chunk := source.read(XLSX_VALIDATION_CHUNK_BYTES):
                        member_uncompressed += len(chunk)
                        total_uncompressed += len(chunk)
                        if member_uncompressed > MAX_XLSX_MEMBER_BYTES:
                            raise BusinessRuleError(
                                "XLSX archive member exceeds the expanded size limit"
                            )
                        if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                            raise BusinessRuleError(
                                "XLSX archive exceeds the total expanded size limit"
                            )

                total_compressed += member.compress_size
                if (
                    member_uncompressed >= MIN_XLSX_RATIO_CHECK_BYTES
                    and member_uncompressed / max(1, member.compress_size)
                    > MAX_XLSX_COMPRESSION_RATIO
                ):
                    raise BusinessRuleError("XLSX archive member compression ratio is unsafe")

            if (
                total_uncompressed >= MIN_XLSX_RATIO_CHECK_BYTES
                and total_uncompressed / max(1, total_compressed) > MAX_XLSX_COMPRESSION_RATIO
            ):
                raise BusinessRuleError("XLSX archive compression ratio is unsafe")
    except BusinessRuleError:
        raise
    except (zipfile.BadZipFile, zipfile.LargeZipFile, OSError, RuntimeError, EOFError) as exc:
        raise BusinessRuleError("XLSX file is not a valid ZIP archive") from exc


def _read_xlsx(content: bytes) -> list[dict[str, Any]]:
    """仅从当前 XLSX 工作表读取已计算的单元格值。"""

    # ZIP 元数据预检必须先于 openpyxl，避免高膨胀内容进入 XML 解析器。
    _validate_xlsx_archive(content)

    # 只读模式限制工作簿内存占用，纯数据模式则防止
    # 公式被解释为可执行的导入表达式。
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise BusinessRuleError("XLSX file could not be parsed") from exc

    # 解析可能在工作表中途失败，因此必须在 `finally` 中关闭工作簿。
    try:
        worksheet = workbook.active
        if worksheet is None or not hasattr(worksheet, "iter_rows"):
            raise BusinessRuleError("the active XLSX sheet must be a worksheet")
        iterator = cast(WorksheetReader, worksheet).iter_rows(values_only=True)
        # 第一行定义与 CSV 相同的字段白名单契约。
        headers = next(iterator, None)
        if headers is None:
            raise BusinessRuleError("import file must contain a header row")
        normalized_headers = [str(value).strip() if value is not None else "" for value in headers]
        _validate_headers(normalized_headers)
        rows: list[dict[str, Any]] = []
        # 每取得一行就检查上限；空行和伪造的巨大行号间隔也计入扫描成本。
        for scanned_rows, values in enumerate(iterator, start=1):
            if scanned_rows > MAX_IMPORT_ROWS:
                raise BusinessRuleError(f"import file exceeds the {MAX_IMPORT_ROWS} row limit")
            if any(value not in (None, "") for value in values):
                rows.append(dict(zip(normalized_headers, values, strict=False)))
        return rows
    finally:
        workbook.close()


def _validate_headers(headers: Sequence[str]) -> None:
    """要求包含规定列，并拒绝无法识别的输入字段。"""

    # 大小写折叠改善表头匹配，同时不改变实际存储值。
    normalized = {header.strip().casefold() for header in headers if header.strip()}
    missing = REQUIRED_COLUMNS - normalized
    unknown = normalized - SUPPORTED_COLUMNS
    # 缺少必需列时数据行无法校验，因此优先报告该问题。
    if missing:
        raise BusinessRuleError(
            f"import file is missing required columns: {', '.join(sorted(missing))}"
        )
    # 拒绝未知列可以暴露拼写错误，而不是静默丢弃数据。
    if unknown:
        raise BusinessRuleError(
            f"import file contains unsupported columns: {', '.join(sorted(unknown))}"
        )


def _normalize_rows(
    rows: list[dict[str, Any]],
    file_hash: str,
) -> list[ParsedTransactionRow]:
    """标准化字段名，并生成支持安全重试插入的稳定标识。"""

    # 遍历前先检查行数，使异常庞大的工作表快速失败。
    if len(rows) > MAX_IMPORT_ROWS:
        raise BusinessRuleError(f"import file exceeds the {MAX_IMPORT_ROWS} row limit")

    # 第一行是表头，因此物理行号从二开始。
    parsed: list[ParsedTransactionRow] = []
    for row_number, row in enumerate(rows, start=2):
        values = {
            str(key).strip().casefold(): value
            for key, value in row.items()
            if key is not None and str(key).strip()
        }
        # 提供方标识符不受文件重排影响；缺少标识符时使用文件哈希
        # 加行位置，确保合法的重复行仍可彼此区分。
        external_id = values.pop("external_id", None)
        identity = (
            f"external:{str(external_id).strip()}"
            if external_id not in (None, "")
            else f"file:{file_hash}:row:{row_number}"
        )
        # 仅持久化摘要，避免泄露提供方标识符。
        import_key = hashlib.sha256(identity.encode("utf-8")).hexdigest()
        parsed.append(
            ParsedTransactionRow(
                row_number=row_number,
                values=values,
                import_key=import_key,
            )
        )
    # 解析后的数据行保持原顺序，以保留按行号报告错误的能力。
    # 校验和数据库变更稍后在服务事务中执行。
    return parsed
