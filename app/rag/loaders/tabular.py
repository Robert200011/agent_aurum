"""Bounded CSV/XLSX parsing with worksheet and source-row coordinates."""

from __future__ import annotations

import csv
import io
import zipfile
from collections.abc import Iterable, Sequence
from datetime import date, datetime, time
from decimal import Decimal

from defusedxml import ElementTree
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.config import Settings
from app.rag.loaders.base import (
    DocumentParsingError,
    ParsedDocument,
    ParsedFragment,
    build_parsed_document,
)
from app.rag.loaders.text import normalize_text

CSV_MIME_TYPE = "text/csv"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_PARSER_VERSION = "aurum-csv-v1"
XLSX_PARSER_VERSION = "aurum-xlsx-openpyxl-v1"


def parse_csv_document(content: bytes, settings: Settings) -> ParsedDocument:
    """Decode UTF-8 CSV and preserve physical source-line ranges."""

    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise DocumentParsingError("CSV document is not valid UTF-8") from exc
    if "\x00" in decoded:
        raise DocumentParsingError("CSV document contains NUL bytes")

    fragments: list[ParsedFragment] = []
    previous_line = 0
    logical_rows = 0
    try:
        reader = csv.reader(io.StringIO(decoded, newline=""), strict=True)
        header: list[str] | None = None
        for raw_row in reader:
            row_start = previous_line + 1
            row_end = reader.line_num
            previous_line = row_end
            row = _bounded_row(raw_row, settings)
            if not any(row):
                continue
            logical_rows += 1
            if logical_rows > settings.document_max_tabular_rows:
                raise DocumentParsingError("CSV document exceeds the configured row limit")
            if header is None:
                header = _normalize_headers(row)
                fragments.append(
                    ParsedFragment(
                        text="Columns: " + " | ".join(header),
                        row_start=row_start,
                        row_end=row_end,
                        metadata={"source_kind": "csv_header"},
                    )
                )
                continue
            fragments.append(
                ParsedFragment(
                    text=_render_row(header, row),
                    row_start=row_start,
                    row_end=row_end,
                    metadata={"source_kind": "csv_row"},
                )
            )
    except csv.Error as exc:
        raise DocumentParsingError("CSV document could not be parsed safely") from exc

    return build_parsed_document(
        fragments=fragments,
        mime_type=CSV_MIME_TYPE,
        parser_version=CSV_PARSER_VERSION,
    )


def parse_xlsx_document(content: bytes, settings: Settings) -> ParsedDocument:
    """Read cached cell values only and preserve worksheet/row coordinates."""

    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (
        ElementTree.ParseError,
        IndexError,
        InvalidFileException,
        KeyError,
        OSError,
        ValueError,
        zipfile.BadZipFile,
    ) as exc:
        raise DocumentParsingError("XLSX document could not be parsed safely") from exc

    fragments: list[ParsedFragment] = []
    logical_rows = 0
    try:
        worksheets = workbook.worksheets
        if len(worksheets) > settings.document_max_workbook_sheets:
            raise DocumentParsingError("XLSX document exceeds the configured worksheet limit")
        for worksheet in worksheets:
            if worksheet.max_column > settings.document_max_tabular_columns:
                raise DocumentParsingError("XLSX worksheet exceeds the configured column limit")
            header: list[str] | None = None
            for row_number, raw_row in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                row = _bounded_row((_format_cell(value) for value in raw_row), settings)
                if not any(row):
                    continue
                logical_rows += 1
                if logical_rows > settings.document_max_tabular_rows:
                    raise DocumentParsingError("XLSX document exceeds the configured row limit")
                if header is None:
                    header = _normalize_headers(row)
                    fragments.append(
                        ParsedFragment(
                            text="Columns: " + " | ".join(header),
                            sheet_name=worksheet.title,
                            row_start=row_number,
                            row_end=row_number,
                            metadata={"source_kind": "xlsx_header"},
                        )
                    )
                    continue
                fragments.append(
                    ParsedFragment(
                        text=_render_row(header, row),
                        sheet_name=worksheet.title,
                        row_start=row_number,
                        row_end=row_number,
                        metadata={"source_kind": "xlsx_row"},
                    )
                )
    except DocumentParsingError:
        raise
    except (
        ElementTree.ParseError,
        IndexError,
        KeyError,
        OSError,
        TypeError,
        ValueError,
        zipfile.BadZipFile,
    ) as exc:
        raise DocumentParsingError("XLSX document could not be parsed safely") from exc
    finally:
        workbook.close()

    return build_parsed_document(
        fragments=fragments,
        mime_type=XLSX_MIME_TYPE,
        parser_version=XLSX_PARSER_VERSION,
    )


def _bounded_row(values: Iterable[object], settings: Settings) -> list[str]:
    row: list[str] = []
    for index, value in enumerate(values):
        if index >= settings.document_max_tabular_columns:
            raise DocumentParsingError("tabular document exceeds the configured column limit")
        normalized = normalize_text(str(value) if value is not None else "").replace("\n", " / ")
        if len(normalized) > settings.document_max_cell_characters:
            raise DocumentParsingError("tabular cell exceeds the configured text limit")
        row.append(normalized)
    while row and not row[-1]:
        row.pop()
    return row


def _normalize_headers(values: Sequence[str]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = value or f"column_{index}"
        count = counts.get(base, 0) + 1
        counts[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _render_row(headers: Sequence[str], values: Sequence[str]) -> str:
    width = max(len(headers), len(values))
    pairs: list[str] = []
    for index in range(width):
        header = headers[index] if index < len(headers) else f"column_{index + 1}"
        value = values[index] if index < len(values) else ""
        if value:
            pairs.append(f"{header}: {value}")
    return " | ".join(pairs) or "(empty row)"


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        return format(value, ".15g")
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)
